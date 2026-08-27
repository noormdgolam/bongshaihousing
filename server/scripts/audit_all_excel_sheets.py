import openpyxl
import json

wb = openpyxl.load_workbook('C:/Users/munna/Downloads/Floor area.XLSX', data_only=True)

print(f'=== AUDITING ALL {len(wb.sheetnames)} SHEETS IN FLOOR AREA.XLSX ===\n')

sheet_analysis = {}

for sname in wb.sheetnames:
    ws = wb[sname]
    rows = []
    for r in range(1, ws.max_row+1):
        vals = [ws.cell(r, c).value for c in range(1, 10)]
        if any(v is not None for v in vals):
            rows.append((r, vals))
    
    print(f'------------------------------------------------------------')
    print(f'SHEET: {sname} (Non-empty rows: {len(rows)})')
    
    # Parse sections, rooms, footprint, subtotal, and total
    current_floor = 'Ground Floor'
    floor_rooms = {current_floor: []}
    footprints = {}
    subtotals = {}
    
    for r_idx, r_vals in rows:
        col0 = str(r_vals[0]).strip() if r_vals[0] is not None else ''
        col1 = r_vals[1] # Area
        col2 = r_vals[2] # Length
        col3 = r_vals[3] # Width
        
        # Check if floor switch
        if 'first floor' in col0.lower() or '1st floor' in col0.lower() or '2nd floor' in col0.lower():
            current_floor = 'First Floor' if ('first' in col0.lower() or '1st' in col0.lower()) else 'Second Floor'
            floor_rooms[current_floor] = []
            continue
        elif 'ground floor' in col0.lower():
            current_floor = 'Ground Floor'
            if current_floor not in floor_rooms:
                floor_rooms[current_floor] = []
            if col1:
                footprints[current_floor] = (col1, col2, col3)
            continue
        
        # Check if header row
        if col0.lower() in ['section', 'overview (area details)', ''] and col1 in ['Area (sqft)', None]:
            continue
        
        # Check if footprint row
        if 'floor area' in col0.lower() or 'plinth area' in col0.lower():
            footprints[current_floor] = (col1, col2, col3)
            continue
        
        # Check if total / subtotal row
        if 'total' in col0.lower() or (col0 == '' and col1 is not None and len([v for v in r_vals if v is not None]) == 1):
            subtotals[current_floor] = col1
            continue
        
        # Regular room
        if col0 and col1 is not None:
            try:
                area_num = float(col1)
                l_num = float(col2) if col2 is not None else None
                w_num = float(col3) if col3 is not None else None
                calc_area = (l_num * w_num) if (l_num and w_num) else None
                
                # Check calculation mismatch
                calc_diff = abs(calc_area - area_num) if calc_area is not None else 0
                calc_flag = f' [MISMATCH: L*W={calc_area} != {area_num}]' if calc_diff > 1 else ''
                
                floor_rooms[current_floor].append({
                    'name': col0,
                    'area': area_num,
                    'l': l_num,
                    'w': w_num,
                    'flag': calc_flag
                })
            except (ValueError, TypeError):
                pass

    # Audit sums for each floor
    grand_sum = 0
    for fl, r_list in floor_rooms.items():
        fl_sum = sum(r['area'] for r in r_list)
        grand_sum += fl_sum
        stated_sub = subtotals.get(fl)
        footprint = footprints.get(fl)
        
        print(f'  [{fl}]: {len(r_list)} rooms, Sum of Rooms = {fl_sum} sqft')
        if stated_sub:
            print(f'    Stated Subtotal in Excel = {stated_sub} (Diff: {fl_sum - float(stated_sub)})')
        if footprint:
            print(f'    Footprint Dimension = {footprint[1]} x {footprint[2]} = {footprint[0]} sqft')
        
        for r in r_list:
            dim_str = f"({r['l']} x {r['w']})" if r['l'] and r['w'] else ""
            print(f"      - {r['name']}: {r['area']} sqft {dim_str}{r['flag']}")

    print(f'  >> GRAND TOTAL (Sum of all rooms): {grand_sum} sqft')
    sheet_analysis[sname] = {
        'rooms_sum': grand_sum,
        'floors': floor_rooms,
        'subtotals': subtotals,
        'footprints': footprints
    }

print('\n=== AUDIT COMPLETE ===')
